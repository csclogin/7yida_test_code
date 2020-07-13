# -*- coding: utf-8 -*-
#  author caishicheng

import pymongo
from openpyxl import load_workbook
from openpyxl import  Workbook


client = pymongo.MongoClient('192.168.198.128',27017)
client = pymongo.MongoClient('dev.7yida.com', 27017, username='admin', password='n7rERJrYQ4DWvCdYdOvzJyS8qmx7Pag5')
db=client.qyd
my_set=db.orders
for i in my_set.find():
    print(i)

workbook = load_workbook(r"C:\Users\PC\Desktop\test.xlsx")
worksheet = workbook.get_sheet_by_name("test")

row=1
cell=1
for i in range(10):
    for j in range(10):
        worksheet.cell(j+1,i+1,2222)
        cell+=1
        row+=1

workbook.save(r"C:\Users\PC\Desktop\test.xlsx")
for row in worksheet.rows:
    print('\n')
    for cell in row:
        print(cell.coordinate,cell.value,end=' ')

for row in worksheet["A1":"F3"]:
    print('\n')
    for cell in row:
        print(cell.coordinate,cell.value,end=' ')